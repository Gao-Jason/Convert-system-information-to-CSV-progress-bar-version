import os
from datetime import datetime
import time
import psutil
import openpyxl
import msvcrt  # For file locking on Windows
from tqdm import tqdm
import sys

def get_desktop_path():
    """Returns the path to the user's desktop."""
    return os.path.join(os.path.expanduser("~"), "Desktop")

def get_latest_date(xlsx_file_path):
    """Returns the latest date from the Excel file."""
    try:
        latest_date = None
        if os.path.exists(xlsx_file_path):
            wb = openpyxl.load_workbook(xlsx_file_path)
            sheet = wb.active
            if sheet.max_row > 1:
                latest_date = sheet.cell(row=sheet.max_row, column=1).value
            wb.close()
        return latest_date
    except Exception as e:
        print(f"Error getting latest date: {e}")
        return None

def get_cpu_info():
    """Returns the average CPU usage percentage."""
    try:
        cpu_percent_per_cpu = psutil.cpu_percent(interval=1, percpu=True)
        avg_cpu_percent = sum(cpu_percent_per_cpu) / len(cpu_percent_per_cpu)
        return [avg_cpu_percent]
    except Exception as e:
        print(f"Error getting CPU info: {e}")
        return None

def get_memory_info():
    """Returns total and available memory in gigabytes."""
    try:
        memory = psutil.virtual_memory()
        memory_total_gb = round(memory.total / (1024 ** 3), 2)
        memory_available_gb = round(memory.available / (1024 ** 3), 2)
        return memory_total_gb, memory_available_gb
    except Exception as e:
        print(f"Error getting memory info: {e}")
        return None, None

def get_disk_info():
    """Returns disk information including free and total space."""
    try:
        disk_info = []
        partitions = psutil.disk_partitions()
        for partition in partitions:
            try:
                usage = psutil.disk_usage(partition.mountpoint)
            except PermissionError:
                continue
            device = partition.device
            freespace_gb = round(usage.free / (1024 * 1024 * 1024), 2)
            size_gb = round(usage.total / (1024 * 1024 * 1024), 2)
            disk_info.append((device, freespace_gb, size_gb))
        return disk_info
    except Exception as e:
        print(f"Error getting disk info: {e}")
        return None

def write_to_excel(cpu_info, memory_info, disk_info, latest_date):
    """Writes system information to an Excel file."""
    try:
        if cpu_info is None or memory_info is None or disk_info is None:
            print("Error: One or more system information is None.")
            return None

        desktop_path = get_desktop_path()
        output_file_path = os.path.join(desktop_path, "system_info.xlsx")

        today_datetime = datetime.now().strftime("%Y-%m-%d,%H:%M")
        today_date = datetime.now().strftime("%Y-%m-%d")

        if latest_date is None:
            latest_date = today_date

        if os.path.exists(output_file_path):
            wb = openpyxl.load_workbook(output_file_path)
            sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["Date", "CPU Usage (%)", "Memory Total (GB)", "Memory Available (GB)", "Disk Free Space (GB)", "Total Disk Space (GB)", "Remarks"])

        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 20

        if disk_info:
            new_row = [today_datetime, cpu_info[0], memory_info[0], memory_info[1], disk_info[0][1], disk_info[0][2], ""]
        else:
            print("Disk info is None.")
            new_row = [today_datetime, cpu_info[0], memory_info[0], memory_info[1], 0, 0, ""]

        sheet.append(new_row)

        try:
            dates = [row[0].split(',')[0] for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=1, values_only=True) if row[0]]
            duplicate_dates = set([x for x in dates if dates.count(x) > 1])
            for date in duplicate_dates:
                duplicate_indices = [i for i, x in enumerate(dates) if x == date]
                for index in duplicate_indices[1:]:
                    row = sheet[index + 2]
                    row[-1].value = "重複"
        except Exception as e:
            print("Error marking duplicate entries:", e)

        wb.save(output_file_path)
        wb.close()

        return output_file_path, today_date  # 返回 output_file_path 和 today_date
    except PermissionError as pe:
        print(f"Permission denied: {pe}")
    except Exception as e:
        print(f"Error writing to Excel file: {e}")
        return None

def main():
    try:
        print("開始執行..請稍等...")
        desktop_path = get_desktop_path()
        output_file_path = os.path.join(desktop_path, "系統資訊.xlsx")

        latest_date = None

        # 获取系统信息
        cpu_info = get_cpu_info()
        memory_info = get_memory_info()
        disk_info = get_disk_info()
        
        # 添加 tqdm
        total_iterations = 1  # 总迭代次数
        for _ in tqdm(range(total_iterations), desc="Getting system info", miniters=20):
            # 将系统信息写入Excel
            result = write_to_excel(cpu_info, memory_info, disk_info, latest_date)
            if result is not None:
                output_file_path, latest_date = result  # 解包元组，更新 output_file_path 和 latest_date

        if output_file_path:
            if os.path.exists(output_file_path):
                print("成功寫入資料到Excel文件。")
            else:
                print("寫入資料到Excel文件失敗。")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        sys.exit()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"An error occurred: {e}")
